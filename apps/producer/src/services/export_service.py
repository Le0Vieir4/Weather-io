"""Export service for CSV and Excel generation"""
import json
import logging
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import EXPORT_PATHS

logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting weather data to files"""
    
    def __init__(self):
        """Initialize export service"""
        self._ensure_export_dirs()
        logger.info("Export service initialized")
    
    def _ensure_export_dirs(self):
        """Ensure export directories exist"""
        os.makedirs(EXPORT_PATHS["csv"], exist_ok=True)
        os.makedirs(EXPORT_PATHS["excel"], exist_ok=True)
    
    def export_csv(self, weather_json):
        """Export weather data to CSV
        
        Args:
            weather_json (str): JSON string with weather data
            
        Returns:
            str: Path to generated CSV file
        """
        try:
            logger.info("Generating CSV export")
            data = json.loads(weather_json)
            df = self._prepare_dataframe(data)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{EXPORT_PATHS['csv']}/weather_data_{timestamp}.csv"
            
            df.to_csv(filename, index=False, encoding='utf-8-sig')
            logger.info(f"CSV exported: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return None
    
    def export_excel(self, weather_json):
        """Export weather data to Excel
        
        Args:
            weather_json (str): JSON string with weather data
            
        Returns:
            str: Path to generated Excel file
        """
        try:
            logger.info("Generating Excel export")
            data = json.loads(weather_json)
            df = self._prepare_dataframe(data)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{EXPORT_PATHS['excel']}/weather_data_{timestamp}.xlsx"
            
            df.to_excel(filename, index=False, engine='openpyxl')
            self._style_excel(filename)
            
            logger.info(f"Excel exported: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            return None
    
    def _prepare_dataframe(self, data):
        """Prepare DataFrame from weather data"""
        location = data.get("location", {})
        current = data.get("current", {})
        daily = data.get("daily", [])
        
        rows = []
        
        # Current weather row
        rows.append({
            "Type": "Current",
            "City": location.get("city", ""),
            "Date/Time": current.get("time", ""),
            "Temperature (°C)": current.get("temperature", ""),
            "Apparent Temperature (°C)": current.get("apparentTemperature", ""),
            "Humidity (%)": current.get("relativeHumidity", ""),
            "UV Index": current.get("uv", ""),
            "Weather": current.get("weatherCode", ""),
            "Precipitation (%)": current.get("precipitationProbability", ""),
            "Max Temp (°C)": "",
            "Min Temp (°C)": ""
        })
        
        # Daily forecast rows
        for day in daily:
            rows.append({
                "Type": "Forecast",
                "City": location.get("city", ""),
                "Date/Time": day.get("date", ""),
                "Temperature (°C)": "",
                "Apparent Temperature (°C)": "",
                "Humidity (%)": "",
                "UV Index": day.get("uvIndexMax", ""),
                "Weather": day.get("weatherCode", ""),
                "Precipitation (%)": day.get("precipitationProbability", ""),
                "Max Temp (°C)": day.get("temperatureMax", ""),
                "Min Temp (°C)": day.get("temperatureMin", "")
            })
        
        return pd.DataFrame(rows)
    
    def _style_excel(self, filename):
        """Apply styling to Excel file"""
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logger.warning(f"Could not apply Excel styling: {e}")


# Backward compatibility functions
def generate_csv():
    """Legacy function for CSV export"""
    from src.services.weather_service import data
    service = ExportService()
    weather_json = data()
    return service.export_csv(weather_json)

def generate_excel():
    """Legacy function for Excel export"""
    from src.services.weather_service import data
    service = ExportService()
    weather_json = data()
    return service.export_excel(weather_json)
